# 必要なライブラリのインポート
from flask import Flask, request, send_file, render_template
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, date, timedelta
import pandas as pd
import tempfile
import os
import io

# Flaskアプリケーションの初期化
# テンプレートフォルダとスタティックフォルダを指定
app = Flask(__name__, template_folder='templates', static_folder='static')
CORS(app)  # CORS（Cross-Origin Resource Sharing）を有効化

@app.route("/upload", methods=["POST"])
def generate_timesheet():
    """
    メインのタイムシート生成機能
    CSV2個とExcel1個のファイルをアップロードして、勤怠データを処理し
    Excelテンプレートに反映して新しいファイルを生成する
    """
    
    # フォームデータの取得
    files = request.files.getlist("files")  # アップロードされたファイルリスト
    name = request.form.get("name")         # 氏名
    eid = request.form.get("eid")           # EID
    organization = request.form.get("organization")  # 組織単位
    year = int(request.form.get("year"))    # 年
    month = int(request.form.get("month"))  # 月
    task = request.form.get("task")         # 業務内容
    president_mode = request.form.get("president") == "on"  # 社長モード（労働時間制限）

    # アップロードファイルの分類と検証
    csv_files = [f for f in files if f.filename.lower().endswith(".csv")]
    if len(csv_files) != 2:
        return "CSV2個をアップロードしてください", 400

    # テンプレートファイルのパスを設定
    template_path = os.path.join("templates", "Excel_templates", "タイムシート(yyyy_mm).xlsx")
    if not os.path.exists(template_path):
        return "テンプレートファイルが見つかりません", 500

    # CSV読み込み＆データ結合処理
    df_all = pd.concat([pd.read_csv(f) for f in csv_files], ignore_index=True)
    
    # 日時データの型変換（エラーの場合はNaTに変換）
    df_all["Work start"] = pd.to_datetime(df_all["Work start"], errors="coerce")
    df_all["Work end"] = pd.to_datetime(df_all["Work end"], errors="coerce")
    df_all["Break start"] = pd.to_datetime(df_all["Break start"], errors="coerce")
    df_all["Break end"] = pd.to_datetime(df_all["Break end"], errors="coerce")
    
    # 日付列を作成し、開始時刻順にソート
    df_all["Date"] = df_all["Work start"].dt.date
    df_all.sort_values("Work start", inplace=True)

    # Excelテンプレートファイルの読み込み
    wb = load_workbook(filename=template_path)
    ws = wb.worksheets[0]  # 最初のワークシートを取得
    ws.title = f"{month}月"  # シート名を設定

    # 月の日数と労働時間制限を設定
    days_in_month = pd.Timestamp(year=year, month=month, day=1).days_in_month
    limit_timedelta = timedelta(hours=4, minutes=48)  # 社長モード用の労働時間制限

    # 基本情報の設定
    d9_date = date(year, month, 1)  # 月初日
    g9_date = date(year, month, days_in_month)  # 月末日
    ws["D6"] = organization # 組織単位
    ws["D8"] = name     # 氏名
    ws["D9"] = d9_date  # 対象月（開始日）
    ws["G9"] = g9_date  # 作業期間終了日（月末）

    # 祝日データの読み込み
    holiday_dict = {}
    try:
        with open("holidays.csv", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line: continue
                parts = line.split(",")
                if len(parts) >= 2:
                    holiday_dict[parts[0]] = parts[1]  # 日付：祝日名の辞書
    except:
        pass  # ファイルが存在しない場合は空辞書のまま

    # 各日のデータ処理ループ
    for day in range(1, days_in_month + 1):
        current_date = date(year, month, day)
        row = 12 + day  # Excelの行番号（12行目から開始）
        date_str = current_date.strftime("%Y-%m-%d")
        day_data = df_all[df_all["Date"] == current_date]  # その日のデータを抽出
        holiday_name = holiday_dict.get(date_str)

        # C列（業務内容・祝日名など）の記入
        if holiday_name:
            ws[f"C{row}"] = holiday_name  # 祝日名を記入
        elif current_date.weekday() < 5:  # 平日（月〜金）
            ws[f"C{row}"] = task  # 業務内容を記入
        else:
            ws[f"C{row}"] = ""  # 土日は空白

        # 勤務時間の処理（勤務データが存在する場合）
        if not day_data.empty:
            # 開始・終了時刻の設定
            start_time = day_data["Work start"].min()
            end_time = day_data["Work end"].max()
            ws[f"H{row}"] = start_time.time()  # 開始時刻
            ws[f"I{row}"] = end_time.time()    # 終了時刻

            # 休憩時間の計算
            valid_breaks = day_data.dropna(subset=["Break start", "Break end"])
            total_break_duration = (valid_breaks["Break end"] - valid_breaks["Break start"]).sum()
            total_work_duration = (end_time - start_time) - total_break_duration

            # 社長モード時の労働時間制限処理
            if president_mode and total_work_duration > limit_timedelta:
                ws[f"K{row}"] = limit_timedelta.total_seconds() / 86400  # 制限時間を記入
                excess = total_work_duration - limit_timedelta
                ws[f"G{row}"] = excess.total_seconds() / 86400  # 超過時間を記入
                ws[f"G{row}"].font = Font(size=12)  # フォントサイズ設定
            else:
                ws[f"K{row}"] = total_work_duration.total_seconds() / 86400  # 実働時間
                ws[f"G{row}"] = None  # 超過時間なし

            # 休憩時間の記入（分単位から日単位に変換）
            break_minutes = int(total_break_duration.total_seconds() // 60)
            ws[f"J{row}"] = break_minutes / 1440  # 1日=1440分

            # 時刻表示フォーマットの設定
            for col in ["H", "I", "J", "K", "G"]:
                if ws[f"{col}{row}"].value is not None:
                    ws[f"{col}{row}"].number_format = "h:mm"

        # 休暇の記入条件（勤務なし・平日・祝日でない）
        elif current_date.weekday() < 5 and not holiday_name:
            ws[f"C{row}"] = "休暇"

    # セル検索用の関数を定義
    def find_cell_by_value(ws, value, column=None):
        """
        指定した値を持つセルを検索する関数
        """
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == value and (column is None or cell.column == column):
                    return cell
        return None

    # 月の日数に応じて不要な日付行を削除
    # 31日まで無い月（28日、29日、30日）の場合、余分な日付行を削除
    if days_in_month < 31:
        # 削除する行の範囲を計算（月末の次の日から31日まで）
        start_delete_row = 12 + days_in_month + 1  # 削除開始行
        end_delete_row = 12 + 31  # 削除終了行（31日分の最後の行）
        
        # 後ろの行から削除（行番号がずれることを防ぐため）
        for row_to_delete in range(end_delete_row, start_delete_row - 1, -1):
            ws.delete_rows(row_to_delete)
        
        # 行削除後にセルの結合を再設定
        # 「実働日」のセル結合を修復
        try:
            # 実働日のセル結合（A列とB列）
            actual_day_cell = find_cell_by_value(ws, "実働日")
            if actual_day_cell:
                merge_range = f"A{actual_day_cell.row}:B{actual_day_cell.row}"
                ws.merge_cells(merge_range)
        except:
            pass  # 結合に失敗しても処理を継続

    # サマリ部分の計算式設定
    end_row = 12 + days_in_month  # データの最終行（行削除後の実際の最終行）

    # 「就業時間」の合計計算式を設定
    work_time_cell = find_cell_by_value(ws, "就業時間", column=5)  # E列で「就業時間」を検索
    if work_time_cell:
        target = ws.cell(row=work_time_cell.row, column=6)  # F列に計算式を設定
        target.value = f"=SUM(K13:K{end_row})/TIME(1,,)"  # 就業時間の合計
        target.number_format = "0.0"  # 数値フォーマット
        
        # 残業時間の計算式を設定（同じ行のI列）
        overtime_target = ws.cell(row=work_time_cell.row, column=9)  # I列に計算式を設定
        overtime_target.value = f"=F{work_time_cell.row}-C{work_time_cell.row}*8"  # 残業時間の計算
        overtime_target.number_format = "0.00"  # 数値フォーマット

    # 「実働日」のカウント計算式を設定
    actual_day_cell = find_cell_by_value(ws, "実働日")
    if actual_day_cell:
        target = ws.cell(row=actual_day_cell.row, column=3)  # C列に計算式を設定
        target.value = f"=COUNTA(H13:H{end_row})"  # 勤務開始時刻が入力された日数をカウント

    # ファイル出力処理
    safe_eid = eid.replace(" ", "_").replace("　", "_")  # ファイル名用に空白が入力されていた場合アンダースコアに変換
    output_filename = f"タイムシート({year:04d}_{month:02d})_{safe_eid}.xlsx"  # 出力ファイル名
    output_stream = io.BytesIO()  # メモリ上のファイルストリーム
    wb.save(output_stream)  # ワークブックをストリームに保存
    output_stream.seek(0)   # ストリームの先頭に移動
    return send_file(output_stream, as_attachment=True, download_name=output_filename)

@app.route("/holidays-ui")
def holidays_ui():
    """
    祝日管理画面を表示するエンドポイント
    """
    return render_template("holidays.html")

@app.route("/holidays/download")
def download_holidays():
    """
    祝日データ（holidays.csv）をダウンロードするエンドポイント
    """
    try:
        return send_file("holidays.csv", as_attachment=True)
    except:
        return "holidays.csv が見つかりません", 404

@app.route("/holidays/upload", methods=["POST"])
def upload_holidays():
    """
    祝日データ（holidays.csv）をアップロードするエンドポイント
    """
    file = request.files.get("file")
    if file and file.filename.endswith(".csv"):
        file.save("holidays.csv")  # アップロードされたファイルを保存
        return "アップロード完了", 200
    return "CSVファイルのみアップロード可能です", 400

# アプリケーションの起動設定
if __name__ == "__main__":
    # 環境変数PORTが設定されていない場合は10000番ポートを使用
    port = int(os.environ.get("PORT", 10000))
    # すべてのインターフェースでアプリケーションを起動
    app.run(host="0.0.0.0", port=port)